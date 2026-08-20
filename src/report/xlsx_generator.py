from io import BytesIO

from datetime import datetime
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import WriteOnlyCell


# Expected report fields order
REPORT_HEADERS = [
    "ticker",
    "name",
    "observations",
    "avg_price",
    "median_price",
    "min_price",
    "max_price",
    "avg_change",
    "avg_change_percent",
    "total_volume",
    "total_value",
]

DEFAULT_FILENAME = "stock_report.xlsx"


# Function for generating XLSX report from analytics
def generate_xlsx_report(
    analytics_df,
    user_email,
    return_buffer=False
):
    if analytics_df is None or analytics_df.empty:
        return None

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet()

    # Header styles
    title_font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="1B4D3E"
    )

    header_fill = PatternFill(
        start_color="1B4D3E",
        end_color="1B4D3E",
        fill_type="solid"
    )

    header_font = Font(
        name="Calibri",
        size=11,
        bold=True,
        color="FFFFFF"
    )

    header_align = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Report title
    title_cell = WriteOnlyCell(sheet, value="Stock Data Service")
    title_cell.font = title_font
    sheet.append([title_cell])

    # Report information
    moscow_time = datetime.now(ZoneInfo("Europe/Moscow"))
    sheet.append(["Инструменты для анализа фондового рынка"])
    sheet.append([])
    sheet.append([f"Экспорт выполнил: {user_email}"])
    sheet.append([f"Дата выгрузки: {moscow_time.strftime('%d.%m.%Y %H:%M')} (MSK)"])
    sheet.append([])

    # Add styled table headers
    styled_headers = []
    for header in REPORT_HEADERS:
        cell = WriteOnlyCell(sheet, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        styled_headers.append(cell)

    sheet.append(styled_headers)

    # Add analytics data
    for _, row in analytics_df.iterrows():
        sheet.append([
                row.get(field, "")
                for field in REPORT_HEADERS
            ]
        )

    if return_buffer:
        buffer = BytesIO()
        workbook.save(buffer)

        # Reset buffer position
        buffer.seek(0)
        return buffer

    # Save report to file
    workbook.save(DEFAULT_FILENAME)

    return DEFAULT_FILENAME
